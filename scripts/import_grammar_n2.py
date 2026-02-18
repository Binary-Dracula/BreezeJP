#!/usr/bin/env python3
"""
从 N2.xlsx 提取语法数据，追加到 SQLite 三表结构：
    grammars → grammar_meanings → grammar_examples

注意: 此脚本在 N5 数据基础上追加 N2 数据（ID 从 N5 最大值 +1 开始）

用法: python3 scripts/import_grammar_n4.py [--dry-run]
"""

import sys
import re
import sqlite3
import shutil
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("需要安装 openpyxl: pip3 install openpyxl")
    sys.exit(1)

DB_PATH = "assets/database/breeze_jp.sqlite"
XLSX_PATH = "files/N2.xlsx"
JLPT_LEVEL = "N2"


def read_all_lines(xlsx_path: str, max_row: int = 0) -> list[str]:
    """读取 xlsx 所有单元格内容，展开多行单元格为独立行"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    
    lines = []
    for idx, row in enumerate(ws.iter_rows(values_only=False)):
        if max_row and idx + 1 > max_row:
            break
        vals = [cell.value for cell in row if cell.value is not None]
        if not vals:
            continue
        cell_text = str(vals[0]).strip()
        if not cell_text:
            continue
        for sub_line in cell_text.split('\n'):
            sub_line = sub_line.strip()
            if sub_line:
                lines.append(sub_line)
    return lines


def normalize_spaced(text: str) -> str:
    """将 '接 续' / '意 思' 等间隔字恢复为正常形式"""
    text = re.sub(r'接\s+续', '接续', text)
    text = re.sub(r'意\s+思', '意思', text)
    text = re.sub(r'提\s+示', '提示', text)
    return text


def parse_lines(lines: list[str]) -> list[dict]:
    """解析扁平化的行列表，返回结构化语法数据"""
    # 标题支持 ~ 和 ～（全角波浪号）
    grammar_title_re = re.compile(r'^(\d+)\s*[\.\．]\s*(.+)$')
    connection_re = re.compile(r'^接续\s*(.*)$')
    meaning_re = re.compile(r'^意思\s*(.*)$')
    numbered_re = re.compile(r'^(\d+)\s*[\.\．]\s*(.+)$')
    example_re = re.compile(r'[◎○]')
    lesson_re = re.compile(r'^第\d+课')
    tip_re = re.compile(r'^提示$')
    exercise_re = re.compile(r'练习|模拟试题|模拟考试')
    
    grammars = []
    current_grammar = None
    current_meaning = None
    in_tip = False
    
    # 找到内容开始
    start_idx = 0
    for i, line in enumerate(lines):
        norm = normalize_spaced(line)
        m = grammar_title_re.match(norm)
        if m and ('~' in m.group(2) or '～' in m.group(2)):
            start_idx = i
            break
        if lesson_re.match(norm):
            start_idx = i
            break
    
    for i in range(start_idx, len(lines)):
        raw = lines[i]
        line = normalize_spaced(raw)
        
        if lesson_re.match(line):
            continue
        
        if current_grammar is not None and not example_re.search(line) and re.match(r'^(第\d+课\s*)?练习|^模拟试题|^模拟考试', line):
            break
        
        # 语法标题
        m = grammar_title_re.match(line)
        if m and ('~' in m.group(2) or '～' in m.group(2)):
            if current_grammar:
                if current_meaning:
                    current_grammar["meanings"].append(current_meaning)
                    current_meaning = None
                grammars.append(current_grammar)
            
            current_grammar = {"title": m.group(2).strip(), "meanings": []}
            in_tip = False
            continue
        
        if not current_grammar:
            continue
        
        # 接续行
        cm = connection_re.match(line)
        if cm:
            if current_meaning:
                current_grammar["meanings"].append(current_meaning)
            
            conn_text = cm.group(1).strip()
            conn_text = re.sub(r'^\(\s*\d+\s*\)\s*', '', conn_text).strip()
            
            current_meaning = {
                "connection": conn_text if conn_text else None,
                "meaning": None,
                "tip": None,
                "examples": [],
                "tip_examples": [],
            }
            in_tip = False
            continue
        
        # 意思行
        mm = meaning_re.match(line)
        if mm:
            meaning_text = mm.group(1).strip()
            if meaning_text:
                if current_meaning is None:
                    current_meaning = {
                        "connection": None,
                        "meaning": meaning_text,
                        "tip": None,
                        "examples": [],
                        "tip_examples": [],
                    }
                else:
                    current_meaning["meaning"] = meaning_text
            continue
        
        # 提示
        if tip_re.match(line):
            in_tip = True
            continue
        
        # 例句
        if example_re.search(line):
            parts = re.split(r'[◎○]', line)
            for part in parts:
                part = part.strip()
                if not part:
                    continue
                sentence, translation = _split_example(part)
                if sentence:
                    ex = {"sentence": sentence, "translation": translation}
                    if current_meaning:
                        if in_tip:
                            current_meaning["tip_examples"].append(ex)
                        else:
                            current_meaning["examples"].append(ex)
            continue
        
        # 编号义项
        nm = numbered_re.match(line)
        if nm and current_meaning is not None:
            number = int(nm.group(1))
            sub_meaning = nm.group(2).strip()
            
            if number > 1:
                current_grammar["meanings"].append(current_meaning)
                current_meaning = {
                    "connection": current_meaning.get("connection"),
                    "meaning": sub_meaning,
                    "tip": None,
                    "examples": [],
                    "tip_examples": [],
                }
            else:
                current_meaning["meaning"] = sub_meaning
            in_tip = False
            continue
        
        # 提示正文
        if in_tip and current_meaning:
            if current_meaning["tip"]:
                current_meaning["tip"] += line
            else:
                current_meaning["tip"] = line
            continue
    
    # 保存最后一条
    if current_grammar:
        if current_meaning:
            current_grammar["meanings"].append(current_meaning)
        grammars.append(current_grammar)
    
    return grammars


def _split_example(text: str) -> tuple[str, str | None]:
    """分离日语例句和中文翻译"""
    text = re.sub(r'\s+', ' ', text).strip()
    parts = text.split('/')
    if len(parts) >= 2:
        for i in range(len(parts) - 1, 0, -1):
            if re.search(r'[\u4e00-\u9fff]', parts[i]):
                sentence = '/'.join(parts[:i]).strip()
                translation = '/'.join(parts[i:]).strip()
                return sentence, translation
    return text, None


def get_next_grammar_id(db_path: str) -> int:
    """获取下一个可用的 grammar ID"""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT MAX(id) FROM grammars")
    max_id = cursor.fetchone()[0]
    conn.close()
    return (max_id or 0) + 1


def import_data(db_path: str, grammars: list[dict], start_id: int, dry_run: bool = False):
    """导入数据（追加模式）"""
    if not dry_run:
        backup_path = f"{db_path}.bak.{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shutil.copy2(db_path, backup_path)
        print(f"数据库备份: {backup_path}")
    
    conn = sqlite3.connect(db_path) if not dry_run else sqlite3.connect(":memory:")
    cursor = conn.cursor()
    now = int(datetime.now().timestamp())
    
    g_count = m_count = e_count = 0
    
    for g_idx, grammar in enumerate(grammars):
        grammar_id = start_id + g_idx
        
        if dry_run:
            print(f"\n[{grammar_id}] {grammar['title']} ({len(grammar['meanings'])} 义项)")
            for m_idx, meaning in enumerate(grammar["meanings"], 1):
                conn_str = meaning.get("connection") or "(无)"
                mean_str = meaning.get("meaning") or "(无)"
                tip_str = " [TIP]" if meaning.get("tip") else ""
                n_ex = len(meaning.get("examples", []))
                n_tip = len(meaning.get("tip_examples", []))
                print(f"  {m_idx}. [{conn_str}] → {mean_str}{tip_str} ({n_ex}+{n_tip}例句)")
                for ex in meaning.get("examples", [])[:2]:
                    s = ex["sentence"][:60]
                    t = (ex.get("translation") or "")[:30]
                    print(f"    ◎ {s} / {t}")
                if len(meaning.get("examples", [])) > 2:
                    print(f"    ... (共{n_ex}条)")
            continue
        
        cursor.execute(
            "INSERT INTO grammars (id, title, jlpt_level, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
            (grammar_id, grammar["title"], JLPT_LEVEL, now, now)
        )
        g_count += 1
        
        for m_idx, meaning in enumerate(grammar["meanings"], 1):
            cursor.execute(
                "INSERT INTO grammar_meanings (grammar_id, sort_order, connection, meaning, tip) VALUES (?, ?, ?, ?, ?)",
                (grammar_id, m_idx, meaning.get("connection"), meaning.get("meaning"), meaning.get("tip"))
            )
            mid = cursor.lastrowid
            m_count += 1
            
            sort = 1
            for ex in meaning.get("examples", []):
                cursor.execute(
                    "INSERT INTO grammar_examples (meaning_id, sort_order, sentence, translation, is_tip_example) VALUES (?, ?, ?, ?, 0)",
                    (mid, sort, ex["sentence"], ex.get("translation"))
                )
                sort += 1
                e_count += 1
            
            for ex in meaning.get("tip_examples", []):
                cursor.execute(
                    "INSERT INTO grammar_examples (meaning_id, sort_order, sentence, translation, is_tip_example) VALUES (?, ?, ?, ?, 1)",
                    (mid, sort, ex["sentence"], ex.get("translation"))
                )
                sort += 1
                e_count += 1
    
    if not dry_run:
        conn.commit()
    conn.close()
    
    prefix = "[DRY RUN] " if dry_run else ""
    print(f"\n{prefix}导入完成: {g_count} 语法 (ID {start_id}-{start_id + len(grammars) - 1}), {m_count} 义项, {e_count} 例句")


def main():
    dry_run = "--dry-run" in sys.argv
    
    if dry_run:
        print("=== DRY RUN 模式 ===\n")
    
    print(f"解析 {XLSX_PATH}...")
    all_lines = read_all_lines(XLSX_PATH, max_row=2724)
    grammars = parse_lines(all_lines)
    # 过滤掉没有义项的条目（误匹配的描述行）
    grammars = [g for g in grammars if len(g["meanings"]) > 0]
    print(f"共解析出 {len(grammars)} 条 {JLPT_LEVEL} 语法\n")
    
    start_id = get_next_grammar_id(DB_PATH)
    print(f"起始 ID: {start_id}")
    
    import_data(DB_PATH, grammars, start_id, dry_run)
    
    if not dry_run:
        print("完成!")


if __name__ == "__main__":
    main()
